import os
import os.path as osp
import sys
from datetime import datetime
from glob import glob
import openpyxl
from openpyxl.styles import PatternFill
import warnings

warnings.filterwarnings("ignore")

def load_file_to_list(filename):
    out_list = []
    with open(filename, 'rb') as f:
        for line in f.readlines():
            out_list.append(line.strip())
    return out_list

def find_key_word(cell, key_words):
    for key_word in key_words:
        if isinstance(cell.value, str) and key_word.decode("utf-8") in cell.value:
            return key_word
    return None

def if_contain_chaos(keyword):
    try:
        keyword.encode("gb2312")
    except UnicodeEncodeError:
        return True
    return False


def process_sheet(sheet, key_words_dict, logger, mode):
    key_words = list(key_words_dict.keys())
    for row in sheet.iter_rows():
        for cell in row:
            try:
                if cell.value is None:
                    continue
                key_word = find_key_word(cell, key_words)
                if key_word is not None:
                    key_words_dict[key_word].append('{}'.format(cell.coordinate))
                    if mode == 'b':
                        cell.value = ''
                        cell.fill = PatternFill("solid", fgColor="000000")
                    else:
                        # replace all the keywords
                        for kw in key_words:
                            cell.value = cell.value.replace(kw.decode("utf-8"), '[DP Redaction]')
            except Exception as ex:
                logger.write('ERROR: cell {} has error {}.\n'.format(cell, ex).encode("utf-8"))

def run_redaction(file_name, key_words, logger, mode):
    # open an excel document
    wb = openpyxl.load_workbook(file_name)
    sheet_dict = {}
    for input_sheet in wb.worksheets:
        print(input_sheet.title)
        key_words_dict = {k:[] for k in key_words}
        process_sheet(input_sheet, key_words_dict, logger, mode)
        sheet_dict[input_sheet.title] = key_words_dict
    file_name, suffix = osp.splitext(osp.basename(file_name))
    wb.save(osp.join('output_files', '{}_redacted{}'.format(file_name, suffix)))

    for key_word in key_words:
        logger.write('* {} \n'.format(key_word.decode("utf-8")).encode("utf-8"))
        for sheet_name, key_words_dict in sheet_dict.items():
            key_words_result = key_words_dict[key_word]
            if len(key_words_result) > 0:
                logger.write('{}: {}\n'.format(sheet_name, ','.join(key_words_result)).encode("utf-8"))
                # logger.write('{}, {}'.format(sheet_name, key_words_result))
                # print('{}, {}'.format(sheet_name, key_words_result))

def trace_sheet(sheet, logger):
    sheet_dict = {
        'dp': [],
        'ss': []
    }

    for row in sheet.iter_rows():
        for cell in row:
            try:
                # from IPython import embed; embed()
                if cell.fill.fill_type != 'solid' or cell.fill.fgColor.index not in [1, '00000000']: # fg is not black
                    continue

                if cell.value is None:
                    sheet_dict['ss'].append('{}'.format(cell.coordinate))
                elif cell.value in ['DP Redaction', 'DP Redacted', 'Redacted']:
                    sheet_dict['dp'].append('{}'.format(cell.coordinate))
                else:
                    pass
            except:
                logger.write('ERROR: cell {} has wrong value.\n'.format(cell).encode("utf-8"))

    return sheet_dict


def trace_file(file_name, logger):
    # open an excel document
    wb = openpyxl.load_workbook(file_name)
    for input_sheet in wb.worksheets:
        sheet_title = input_sheet.title
        print(sheet_title)
        sheet_dict = trace_sheet(input_sheet, logger)
        logger.write('* {} \n'.format(sheet_title).encode("utf-8"))
        for op_name, op_coord in sheet_dict.items():
            if len(op_coord) > 0:
                logger.write('{}: {}\n'.format(op_name, ','.join(op_coord)).encode("utf-8"))
                logger.flush()


def run_keyword(input_files):

    if not osp.exists('input_files/keywords.txt'):
        print('Keywords.txt can not be found in input_files directory.')
        input("Press Enter to Quit...")
        sys.exit(0)

    key_words = load_file_to_list('input_files/keywords.txt')
    print('Some chaos keywords are listed below:')
    is_chaos = False
    for key_word_index, key_word in enumerate(key_words):
        if if_contain_chaos(key_word.decode("utf-8")):
            is_chaos = True
            print('Keyword {}:{}'.format(key_word_index, key_word.decode("utf-8")))
    if is_chaos:
        is_continue = input("[c]ontinue or [s]top to correct the above keywords.")
        key_word_correct = 'n' if is_continue == 's' else 'y'
    else:
        input("No chaos keywords are found. Please press enter to continue.")
        key_word_correct = 'y'

    if key_word_correct is 'y':
        mode = input("Use [r]eplace or [b]lack?\n")
        print('You choose {}'.format(mode))
        while mode not in ['r', 'b']:
            mode = input('Please input r or b. r -> repalce, b -> black.')
        # check output directory
        if not osp.exists('output_files'):
            os.mkdir('output_files')
        current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
        logger = open('output_files/log-{}.txt'.format(current_time), 'wb')

        # start process files
        print('Hi, I am processing your files one by one...')
        for input_file in input_files:
            print(input_file)
            logger.write('=========== {} ===========\n'.format(input_file).encode("utf-8"))
            run_redaction(input_file, key_words, logger, mode)
            logger.write('\n'.encode("utf-8"))
            logger.flush()
        logger.close()
        print('All files have been processed.')
    else:
        print('Please modify keywords.txt.')

def run_trace(input_files):
    current_time = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    logger = open('output_files/log-{}.txt'.format(current_time), 'wb')

    # start process files
    print('Hi, I am processing your files one by one...')
    for input_file in input_files:
        print(input_file)
        logger.write('=========== {} ===========\n'.format(input_file).encode("utf-8"))
        trace_file(input_file, logger)
        logger.write('\n'.encode("utf-8"))
        logger.flush()
    logger.close()
    print('All files have been processed.')


if __name__ == '__main__':

    print('''
                  _                           _            _   _             
       __ _ _   _| |_ ___        _ __ ___  __| | __ _  ___| |_(_) ___  _ __  
      / _` | | | | __/ _ \ _____| '__/ _ \/ _` |/ _` |/ __| __| |/ _ \| '_ \ 
     | (_| | |_| | || (_) |_____| | |  __/ (_| | (_| | (__| |_| | (_) | | | |
      \__,_|\__,_|\__\___/      |_|  \___|\__,_|\__,_|\___|\__|_|\___/|_| |_|
                                                                             
    ''')

    # check input xlsx files
    input_files = glob('input_files/*.xlsx')
    if len(input_files) == 0:
        print('No xlsx files are found in input_files directory.')
        input("Press Enter to Quit...")
        sys.exit(0)

    mode = input("[R]eplace or [T]race ?")
    if mode.lower() == 'r':
        run_keyword(input_files)
    else:
        run_trace(input_files)

    input("Press Enter to Quit...")
    # from IPython import embed; embed()